from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, PatternFill

root = Path('/Users/doug/Development/Node/CodexCharts')
out_path = root / 'chart-contract' / 'templates' / 'line-series-template.xlsx'
max_template_rows = 5000

wb = Workbook()
ws = wb.active
ws.title = 'Data'

# Wide-series headers support 2 or 3 line series in one chart.
# A=date, B=primary series, C=secondary series, D=optional third series.
headers = ['date', 'WaterYear', 'Average', 'OptionalSeries']
ws.append(headers)

header_fill = PatternFill(fill_type='solid', fgColor='184F3F')
header_font = Font(color='FFFFFF', bold=True)
for col_idx, title in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col_idx)
    c.value = title
    c.fill = header_fill
    c.font = header_font

# Seed sample rows so the chart renders immediately
sample_rows = [
    [date(2025, 1, 1), 2200, 2000, None],
    [date(2025, 2, 1), 2400, 2050, None],
    [date(2025, 3, 1), 2100, 2150, None],
    [date(2025, 4, 1), 2600, 2250, None],
]
for row in sample_rows:
    ws.append(row)

for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=1).number_format = 'mmm'

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 16

chart = LineChart()
chart.title = 'Line Series Template (Edit Data sheet to update)'
chart.style = 10
chart.y_axis.title = 'Flow (cfs)'
chart.height = 10
chart.width = 22

date_axis = DateAxis()
date_axis.number_format = 'mmm'
date_axis.majorTimeUnit = 'months'
date_axis.title = 'Month'
chart.x_axis = date_axis

# Chart uses B:D as series and A as category axis.
# Keep chart references on a fixed row window so template-filling can avoid mutating chart objects.
data_ref = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=max_template_rows)
cat_ref = Reference(ws, min_col=1, min_row=2, max_row=max_template_rows)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cat_ref)
chart.legend.position = 't'

ws.add_chart(chart, 'F2')

meta = wb.create_sheet('TemplateMeta')
meta['A1'] = 'Template purpose'
meta['B1'] = 'Editable line-series export with embedded chart'
meta['A2'] = 'Required columns'
meta['B2'] = 'A=date, B=first series, C=second series'
meta['A3'] = 'Optional columns'
meta['B3'] = 'D=third series (optional); D may be blank'
meta['A4'] = 'How to use'
meta['B4'] = 'Replace or append rows in Data sheet; chart updates automatically.'
meta['A5'] = 'Chart anchor'
meta['B5'] = 'Data!F2'
meta['A6'] = 'Generator'
meta['B6'] = 'scripts/generate_excel_template.py'
meta['A7'] = 'Chart data row window'
meta['B7'] = f'Rows 2..{max_template_rows} on Data sheet'
meta.column_dimensions['A'].width = 20
meta.column_dimensions['B'].width = 72

wb.save(out_path)
print(f'Created template: {out_path}')
